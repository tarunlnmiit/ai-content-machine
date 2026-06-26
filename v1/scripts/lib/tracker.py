#!/usr/bin/env python3
"""Annual tracker integration — read/write output/trackers/annual-tracker-2026.xlsx.

The tracker is the authoritative angle-dedup source (CLAUDE.md: "TRACKER FIRST").
It has one sheet per month (May–Dec), with these columns:
  ISO Week · Slug · Day · Date · Posting Date · Time · Niche · Platform · Format · Content Title · Status · ✓

Public API
----------
read_recent_titles(niche, days=90) -> list[str]
    Returns Content Titles posted (Posting Date) within the last `days` days for this niche.
    Used by produce_blog.py to avoid repeating angles. Returns [] if tracker absent.

mark_published(slug, title=None, status="Published") -> int
    Finds rows matching the slug (or title fallback) and sets their Status column.
    Returns the count of rows updated. No-op if tracker absent.

add_row(row_data) -> None
    Appends a new row to the appropriate month sheet (derived from Posting Date).
    Call after repurpose_blog.py generates derivatives to register new content.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

REPO = Path(__file__).resolve().parent.parent.parent
TRACKER_PATH = REPO / "output" / "trackers" / "annual-tracker-2026.xlsx"

# Map tracker Niche column → internal niche key
_NICHE_FROM_TRACKER: dict[str, str] = {
    "DS":      "ds",
    "Life":    "life",
    "Poetry":  "poetry",
}
# Reverse map: internal → tracker
_NICHE_TO_TRACKER: dict[str, str] = {v: k for k, v in _NICHE_FROM_TRACKER.items()}

# Column order (0-indexed) — mirrors the header row in the xlsx.
_COL = {
    "ISO Week":      0,
    "Slug":          1,
    "Day":           2,
    "Date":          3,
    "Posting Date":  4,
    "Time":          5,
    "Niche":         6,
    "Platform":      7,
    "Format":        8,
    "Content Title": 9,
    "Status":        10,
    "✓":             11,
}

_DATE_FMTS = ("%d %b %Y", "%d %B %Y", "%-d %b %Y", "%-d %B %Y")


def _parse_tracker_date(val: object) -> Optional[date]:
    """Parse a Posting Date cell value → date object. Returns None on failure."""
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s == "—":
        return None
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Last resort: try ISO
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _open_wb():
    """Import openpyxl and return the workbook (or None if tracker missing/unreadable)."""
    if not TRACKER_PATH.exists():
        return None
    try:
        import openpyxl
        return openpyxl.load_workbook(str(TRACKER_PATH))
    except Exception:
        return None


def read_recent_titles(niche: str, days: int = 90) -> list[str]:
    """Return Content Titles from the tracker posted within the last `days` days for `niche`.

    Deduplicated by slug (same slug can appear across many platform rows — count it once).
    Returns [] if tracker is absent or unreadable — caller falls back to filesystem scan.
    """
    wb = _open_wb()
    if wb is None:
        return []

    tracker_niche = _NICHE_TO_TRACKER.get(niche)
    if not tracker_niche:
        return []

    cutoff = date.today() - timedelta(days=days)
    seen_slugs: set[str] = set()
    titles: list[str] = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            niche_val = row[_COL["Niche"]]
            if niche_val != tracker_niche:
                continue
            posting_date = _parse_tracker_date(row[_COL["Posting Date"]])
            if posting_date is None or posting_date < cutoff:
                continue
            title = row[_COL["Content Title"]]
            if not title:
                continue
            # Deduplicate: use slug if present, else title string
            slug_val = row[_COL["Slug"]] or str(title)
            if slug_val in seen_slugs:
                continue
            seen_slugs.add(slug_val)
            titles.append(str(title))

    return titles


def mark_published(slug: str, title: Optional[str] = None,
                   status: str = "Published") -> int:
    """Update Status to `status` for all rows matching `slug` (or `title` fallback).

    Matches on the Slug column first; if slug is None in a row but title matches,
    that row is updated too. Saves the workbook in-place.
    Returns the number of rows updated.
    """
    try:
        import openpyxl
    except ImportError:
        return 0
    if not TRACKER_PATH.exists():
        return 0
    try:
        wb = openpyxl.load_workbook(str(TRACKER_PATH))
    except Exception:
        return 0

    updated = 0
    status_col_idx = _COL["Status"] + 1   # openpyxl is 1-indexed

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            slug_cell  = row[_COL["Slug"]]
            title_cell = row[_COL["Content Title"]]
            status_cell = row[_COL["Status"]]

            slug_match  = slug_cell.value and slug_cell.value == slug
            title_match = (
                title is not None
                and title_cell.value
                and str(title_cell.value).strip() == title.strip()
            )

            if slug_match or title_match:
                status_cell.value = status
                updated += 1

    if updated:
        wb.save(str(TRACKER_PATH))

    return updated


def tracker_niche_label(niche: str) -> str:
    """Convert internal niche key ('ds') to tracker label ('DS')."""
    return _NICHE_TO_TRACKER.get(niche, niche)
